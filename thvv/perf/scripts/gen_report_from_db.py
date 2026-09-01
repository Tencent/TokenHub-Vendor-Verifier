"""Generate performance report from benchmark_data.db.

Usage:
  python3 scripts/gen_report_from_db.py                         # 自动扫描 results/ 下所有压测目录
  python3 scripts/gen_report_from_db.py --client "美国硅谷"    # 指定客户端位置
  python3 scripts/gen_report_from_db.py --out report.xlsx       # 指定输出文件名

输出：results/性能测试报告_<时间戳>.xlsx

Sheets:
  - 性能指标       38 列聚合指标（每行一个 run）

失败请求的逐条明细不再放 xlsx（与 Excel 指标交付版定位解耦），
统一由 性能测试报告.html（gen_perf_dashboard.py 第四章「失败请求明细」）承载。
"""
import argparse
import json
import math
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


_VENDOR_CHART_JS = Path(__file__).resolve().parent.parent.parent / "scripts" / "vendor" / "chart.umd.min.js"


def _chart_js_inline_or_cdn() -> str:
    """优先内联本地 vendor/chart.umd.min.js，失败则回退 CDN。"""
    try:
        if _VENDOR_CHART_JS.is_file():
            return f"<script>{_VENDOR_CHART_JS.read_text(encoding='utf-8')}</script>"
    except Exception:
        pass
    return '<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>'


def percentile(data, p):
    if not data:
        return None
    s = sorted(data)
    k = (len(s) - 1) * p / 100
    f, c = math.floor(k), math.ceil(k)
    if f == c:
        return s[f]
    return s[f] + (s[c] - s[f]) * (k - f)


def safe_avg(data):
    return sum(data) / len(data) if data else None


def _window_midpoints(t0: float, t_end: float, window_sec: float, step_sec: float | None = None) -> list[tuple[float, float, float]]:
    """生成滑动窗口 (lo, hi, mid_relative)。"""
    if t0 is None or t_end is None or t_end <= t0:
        return []
    step_sec = step_sec or window_sec / 2
    windows = []
    lo = t0
    while lo <= t_end:
        hi = lo + window_sec
        mid = lo + window_sec / 2 - t0
        windows.append((lo, hi, round(mid, 1)))
        lo += step_sec
    return windows


def _point_rate_trends(events: list[tuple[float, float]], t0: float, t_end: float, window_sec: float = 60.0) -> list[tuple[float, float]]:
    """Counter+rate 的离线近似：点事件按滑动窗口聚合为每分钟速率。"""
    import bisect
    events = sorted((float(ts), float(v)) for ts, v in events if ts is not None and v is not None)
    if not events:
        return []
    times = [ts for ts, _ in events]
    prefix_sum = [0.0]
    for _, value in events:
        prefix_sum.append(prefix_sum[-1] + value)

    def _range_sum(lo, hi):
        left = bisect.bisect_left(times, lo)
        right = bisect.bisect_left(times, hi)
        return prefix_sum[right] - prefix_sum[left]

    window_min = window_sec / 60.0
    trends = []
    for lo, hi, mid in _window_midpoints(t0, t_end, window_sec):
        trends.append((mid, round(_range_sum(lo, hi) / window_min, 1)))
    return trends


def _interval_rate_trends(intervals: list[tuple[float, float, float]], t0: float, t_end: float, window_sec: float = 60.0) -> list[tuple[float, float]]:
    """把区间内持续产生的 token 按重叠时间摊分到滑动窗口，近似 output token rate。"""
    norm = []
    for start, end, tokens in intervals:
        if start is None or end is None or tokens is None:
            continue
        start, end, tokens = float(start), float(end), float(tokens)
        if tokens <= 0:
            continue
        if end <= start:
            end = start + 1e-6
        norm.append((start, end, tokens / max(end - start, 1e-6)))
    if not norm:
        return []

    window_min = window_sec / 60.0
    trends = []
    for lo, hi, mid in _window_midpoints(t0, t_end, window_sec):
        token_sum = 0.0
        for start, end, rate_per_sec in norm:
            overlap = max(0.0, min(hi, end) - max(lo, start))
            if overlap > 0:
                token_sum += rate_per_sec * overlap
        trends.append((mid, round(token_sum / window_min, 1)))
    return trends


def _merge_rate_trends(*trend_lists: list[tuple[float, float]]) -> list[tuple[float, float]]:
    """按时间点合并多条 rate 曲线。"""
    merged: dict[float, float] = {}
    for trends in trend_lists:
        for t, v in trends or []:
            key = round(float(t), 1)
            merged[key] = merged.get(key, 0.0) + (float(v) if v is not None else 0.0)
    return [(t, round(v, 1)) for t, v in sorted(merged.items())]


def _compute_token_trends(records: list, window_sec: float = 60.0) -> dict:
    """按业界 Counter+rate 口径离线近似 Input/Output/Total TPM 与 QPM。

    - Input TPM：prompt_tokens 按 start_time 归桶，近似请求注入/prefill 压力。
    - Output TPM：completion_tokens 按 [start_time + TTFT, completed_time] 区间均摊，
      比全部归到 completed_time 更接近流式生成阶段的真实吞吐。
    - Total TPM：Input TPM + Output TPM。
    - QPM：成功请求按 completed_time 归桶，近似 request_success_total 的 rate。
    """
    valid = [r for r in records if r.get("start_time") is not None and r.get("completed_time") is not None]
    if not valid:
        return {"input": [], "output": [], "total": [], "qpm": []}

    t0 = min(float(r["start_time"]) for r in valid)
    t_end = max(float(r["completed_time"]) for r in valid)
    if t_end <= t0:
        return {"input": [], "output": [], "total": [], "qpm": []}

    input_events = []
    output_intervals = []
    qpm_events = []
    for r in valid:
        st = float(r["start_time"])
        ct = float(r["completed_time"])
        prompt_tokens = float(r.get("prompt_tokens") or 0)
        completion_tokens = float(r.get("completion_tokens") or 0)
        if prompt_tokens > 0:
            input_events.append((st, prompt_tokens))
        if completion_tokens > 0:
            ttft = r.get("first_chunk_latency")
            output_start = st + float(ttft) if isinstance(ttft, (int, float)) and ttft >= 0 else st
            output_start = min(max(output_start, st), ct)
            output_intervals.append((output_start, ct, completion_tokens))
        qpm_events.append((ct, 1.0))

    input_trends = _point_rate_trends(input_events, t0, t_end, window_sec)
    output_trends = _interval_rate_trends(output_intervals, t0, t_end, window_sec)
    total_trends = _merge_rate_trends(input_trends, output_trends)
    qpm_trends = _point_rate_trends(qpm_events, t0, t_end, window_sec)
    return {"input": input_trends, "output": output_trends, "total": total_trends, "qpm": qpm_trends}


def calc_stats(db_path, label, parallel, tpm_window_sec=60.0):
    """从 benchmark_data.db 计算性能指标 + TPM/QPM 趋势。"""
    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(result)")
    cols = [r[1] for r in cur.fetchall()]
    cur.execute("SELECT * FROM result")
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()

    success = [r for r in rows if r.get("success") == 1]
    failed = [r for r in rows if r.get("success") != 1]
    total = len(rows)

    if not success:
        return None

    ttfts = [r["first_chunk_latency"] for r in success if r.get("first_chunk_latency") is not None]
    ttlts = [r["latency"] for r in success if r.get("latency") is not None]

    rates = []
    for r in success:
        ct = r.get("completion_tokens", 0) or 0
        lat = r.get("latency") or 0
        ttft = r.get("first_chunk_latency") or 0
        gen_t = lat - ttft
        if ct >= 5 and gen_t >= 0.1:
            rates.append(ct / gen_t)

    itls = []
    for r in success:
        itl_str = r.get("inter_token_latencies")
        if itl_str:
            try:
                vals = json.loads(itl_str) if isinstance(itl_str, str) else itl_str
                if isinstance(vals, list):
                    itls.extend(v for v in vals if isinstance(v, (int, float)))
            except (json.JSONDecodeError, TypeError):
                pass

    prompt_tokens = [r.get("prompt_tokens", 0) or 0 for r in success]
    completion_tokens = [r.get("completion_tokens", 0) or 0 for r in success]
    total_tokens_sum = sum(prompt_tokens) + sum(completion_tokens)

    start_times = [r["start_time"] for r in success]
    completed_times = [r["completed_time"] for r in success]
    total_time_s = max(completed_times) - min(start_times)

    output_tps = sum(completion_tokens) / total_time_s if total_time_s > 0 else 0
    output_tpm = output_tps * 60
    input_tps = sum(prompt_tokens) / total_time_s if total_time_s > 0 else 0
    input_tpm = input_tps * 60
    tpm = input_tpm + output_tpm

    # TPM/QPM 滑动窗口趋势：按业界 Counter+rate 口径离线近似（默认 60s 窗口，50% 重叠）
    token_trends = _compute_token_trends(success, window_sec=float(tpm_window_sec)) if success else {
        "input": [], "output": [], "total": [], "qpm": []
    }

    return {
        "数据集": label,
        "并发数": parallel,
        "总请求数": total,
        "成功数": len(success),
        "失败数": len(failed),
        "成功率": len(success) / total if total > 0 else 0,
        "Avg TTFT(s)": safe_avg(ttfts),
        "Min TTFT(s)": min(ttfts) if ttfts else None,
        "Max TTFT(s)": max(ttfts) if ttfts else None,
        "P50 TTFT(s)": percentile(ttfts, 50),
        "P90 TTFT(s)": percentile(ttfts, 90),
        "P95 TTFT(s)": percentile(ttfts, 95),
        "P99 TTFT(s)": percentile(ttfts, 99),
        "Avg TTLT(s)": safe_avg(ttlts),
        "P50 TTLT(s)": percentile(ttlts, 50),
        "P90 TTLT(s)": percentile(ttlts, 90),
        "P95 TTLT(s)": percentile(ttlts, 95),
        "P99 TTLT(s)": percentile(ttlts, 99),
        "Avg Rate(t/s)": safe_avg(rates),
        "P50 Rate(t/s)": percentile(rates, 50),
        "P90 Rate(t/s)": percentile(rates, 90),
        "P95 Rate(t/s)": percentile(rates, 95),
        "P99 Rate(t/s)": percentile(rates, 99),
        "Avg ITL(s)": safe_avg(itls),
        "P50 ITL(s)": percentile(itls, 50),
        "P90 ITL(s)": percentile(itls, 90),
        "P95 ITL(s)": percentile(itls, 95),
        "P99 ITL(s)": percentile(itls, 99),
        "Avg Prompt Tokens": safe_avg(prompt_tokens),
        "Avg Completion Tokens": safe_avg(completion_tokens),
        "Total Tokens": total_tokens_sum,
        "Avg Cached Tokens": 0,
        "Avg Cache Hit Ratio": 0,
        "Avg Total Time(ms)": (safe_avg(ttlts) or 0) * 1000,
        "Output TPM": output_tpm,
        "Output TPS": output_tps,
        "Input TPM": input_tpm,
        "TPM": tpm,
        # 内部用，不写入「性能指标」sheet
        "_input_tpm_trends": token_trends["input"],
        "_output_tpm_trends": token_trends["output"],
        "_total_tpm_trends": token_trends["total"],
        "_tpm_trends": token_trends["total"],  # 兼容旧 HTML 逻辑
        "_qpm_trends": token_trends["qpm"],
        "_parallel": parallel,
    }


HEADERS = [
    "客户端", "后端网路链接",
    "数据集", "并发数", "总请求数", "成功数", "失败数", "成功率",
    "Avg TTFT(s)", "Min TTFT(s)", "Max TTFT(s)",
    "P50 TTFT(s)", "P90 TTFT(s)", "P95 TTFT(s)", "P99 TTFT(s)",
    "Avg TTLT(s)", "P50 TTLT(s)", "P90 TTLT(s)", "P95 TTLT(s)", "P99 TTLT(s)",
    "Avg Rate(t/s)", "P50 Rate(t/s)", "P90 Rate(t/s)", "P95 Rate(t/s)", "P99 Rate(t/s)",
    "Avg ITL(s)", "P50 ITL(s)", "P90 ITL(s)", "P95 ITL(s)", "P99 ITL(s)",
    "Avg Prompt Tokens", "Avg Completion Tokens", "Total Tokens",
    "Avg Cached Tokens", "Avg Cache Hit Ratio",
    "Avg Total Time(ms)",
    "Output TPM", "Output TPS", "Input TPM", "TPM",
]


def find_run_dirs(results_root: Path):
    """扫描 results/ 下含 benchmark_data.db 的子目录，返回 (run_dir, db_path, vendor, bucket, parallel)"""
    runs = []
    for db in results_root.rglob("benchmark_data.db"):
        # db 路径形如 results/<run-tag>/<model>/parallel_<P>_number_<N>/benchmark_data.db
        try:
            parts = db.relative_to(results_root).parts
            run_tag = parts[0]                      # 例如 "tokenhub-1k-20260519-1008-r2"
            parallel_dir = parts[-2]                # 例如 "parallel_10_number_500"
            m = re.match(r"parallel_(\d+)_number_(\d+)", parallel_dir)
            parallel = int(m.group(1)) if m else 0
            # 解析 vendor 和 bucket（约定：<vendor>-<bucket>-<timestamp>...）
            tag_parts = run_tag.split("-")
            vendor = tag_parts[0]
            # 找第一个像 1k/9k/16k/50k 的 token
            bucket = next((p for p in tag_parts if re.fullmatch(r"\d+k", p)), "?")
            runs.append((run_tag, db, vendor, bucket, parallel))
        except Exception:
            continue
    runs.sort(key=lambda r: r[0])
    return runs


def main():
    ap = argparse.ArgumentParser(description="从 benchmark_data.db 直接生成 38 列性能报告")
    ap.add_argument("--results-dir", default="results", help="results 目录（默认 ./results）")
    ap.add_argument("--client", default="", help="客户端位置（如：美国硅谷 / 国内广州）")
    ap.add_argument("--network", default="", help="后端网络链接说明")
    ap.add_argument("--out", default="", help="输出 xlsx 路径，默认 results/性能测试报告_<ts>.xlsx")
    ap.add_argument("--filter", default="", help="只匹配 run-tag 包含该字符串的实验")
    ap.add_argument("--tpm-window-sec", type=float, default=60.0, help="TPM/QPM 趋势图滑动窗口秒数，默认 60")
    ap.add_argument("--no-tpm-html", action="store_true", help="不生成 TPM 趋势 HTML（默认与 xlsx 同名伴随生成）")
    args = ap.parse_args()

    root = Path(args.results_dir)
    runs = find_run_dirs(root)
    if args.filter:
        runs = [r for r in runs if args.filter in r[0]]
    if not runs:
        print(f"[warn] {root} 下未找到任何 benchmark_data.db", file=sys.stderr)
        sys.exit(1)

    rows = []
    for run_tag, db, vendor, bucket, parallel in runs:
        stats = calc_stats(db, bucket, parallel, args.tpm_window_sec)
        if not stats:
            print(f"[skip] {run_tag} 无成功样本")
            continue
        stats["客户端"] = args.client
        stats["后端网路链接"] = args.network or vendor
        stats["数据集"] = f"{bucket} ({run_tag})"
        rows.append(stats)
        print(f"[{run_tag}] success={stats['成功数']}/{stats['总请求数']} | "
              f"TTFT={stats['Avg TTFT(s)']:.3f}s | TTLT={stats['Avg TTLT(s)']:.3f}s | "
              f"OutTPS={stats['Output TPS']:.2f}")

    if not rows:
        print("[warn] 没有可用的 row", file=sys.stderr)
        sys.exit(1)

    # ---- 写 xlsx ----
    wb = Workbook()
    ws = wb.active
    ws.title = "性能指标"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for ridx, r in enumerate(rows, start=2):
        for c, h in enumerate(HEADERS, start=1):
            v = r.get(h)
            cell = ws.cell(row=ridx, column=c, value=v)
            if h.startswith(("Avg ", "Min ", "Max ", "P50 ", "P90 ", "P95 ", "P99 ")):
                if "(s)" in h or "(t/s)" in h:
                    cell.number_format = "0.000"
                elif "(ms)" in h:
                    cell.number_format = "0.00"
                elif "Tokens" in h:
                    cell.number_format = "0.0"
            elif h in ("成功率", "Avg Cache Hit Ratio"):
                if isinstance(v, (int, float)):
                    cell.number_format = "0.00%" if 0 <= v <= 1 else "0.00"
            elif h == "Total Tokens":
                cell.number_format = "0"
            elif h in ("Output TPM", "Output TPS", "Input TPM", "TPM"):
                cell.number_format = "0.00"
            elif h in ("总请求数", "成功数", "失败数", "并发数"):
                cell.number_format = "0"

    widths = {"数据集": 30, "并发数": 8, "总请求数": 10, "成功数": 8, "失败数": 8, "成功率": 9}
    for c, h in enumerate(HEADERS, start=1):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = widths.get(h, 14)
    ws.freeze_panes = "B2"

    out_path = Path(args.out) if args.out else (
        root / f"性能测试报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n[OK] Saved: {out_path}")
    print(f"     性能指标: {len(rows)} rows x {len(HEADERS)} cols")

    # ── TPM/QPM HTML 趋势图 ──
    if not args.no_tpm_html:
        write_tpm_html(out_path, rows)


def write_tpm_html(xlsx_path, rows):
    # 基于 Chart.js 生成 Input/Output/Total TPM & QPM 趋势 HTML，与 xlsx 同目录。
    if not rows:
        return
    from collections import defaultdict

    groups = defaultdict(list)
    for r in rows:
        label = r.get("数据集", "?")
        groups[label].append(r)

    palette_str = json.dumps([
        {"border": "#2563eb", "bg": "rgba(37,99,235,.10)"},
        {"border": "#dc2626", "bg": "rgba(220,38,38,.10)"},
        {"border": "#059669", "bg": "rgba(5,150,105,.10)"},
        {"border": "#d97706", "bg": "rgba(217,119,6,.10)"},
        {"border": "#7c3aed", "bg": "rgba(124,58,237,.10)"},
        {"border": "#0891b2", "bg": "rgba(8,145,178,.10)"},
        {"border": "#be185d", "bg": "rgba(190,24,93,.10)"},
    ], ensure_ascii=False)

    js_datasets = {}
    metric_defs = [
        ("total", "Total TPM"),
        ("input", "Input TPM"),
        ("output", "Output TPM"),
        ("qpm", "QPM"),
    ]
    for label, group in sorted(groups.items()):
        parallels = sorted(set(r.get("_parallel", 0) for r in group))
        trends_by_metric = {key: {} for key, _ in metric_defs}
        all_times = set()
        for r in group:
            p = r.get("_parallel", 0)
            trends_by_metric["total"][p] = r.get("_total_tpm_trends") or r.get("_tpm_trends", [])
            trends_by_metric["input"][p] = r.get("_input_tpm_trends", [])
            trends_by_metric["output"][p] = r.get("_output_tpm_trends", [])
            trends_by_metric["qpm"][p] = r.get("_qpm_trends", [])
            for metric_map in trends_by_metric.values():
                for t, _ in metric_map.get(p, []) or []:
                    all_times.add(round(t, 1))

        time_axis = sorted(all_times)

        def _lookup(trends, t):
            for tt, tv in trends or []:
                if abs(tt - t) < 0.1:
                    return tv
            return None

        rows_out = []
        cols = {}
        for metric_key, metric_title in metric_defs:
            cols[metric_key] = [f"P{p} {metric_title}" for p in parallels]
        for t in time_axis:
            rec = {"时间(s)": t}
            for metric_key, metric_title in metric_defs:
                for p in parallels:
                    rec[f"P{p} {metric_title}"] = _lookup(trends_by_metric[metric_key].get(p, []), t)
            rows_out.append(rec)
        js_datasets[label] = {"rows": rows_out, "cols": cols}

    labels = list(js_datasets.keys())
    if not labels:
        return
    first_label = labels[0]
    html_path = str(xlsx_path).rsplit(".", 1)[0] + "_TPM.html"
    html = _HTML_TPM_TEMPLATE.format(
        labels_json=json.dumps(js_datasets, ensure_ascii=False),
        first_label=first_label,
        palette=palette_str,
        options_html="\n".join(f'<option value="{l}">{l}</option>' for l in labels),
    )
    html = html.replace("__CHART_JS_TAG__", _chart_js_inline_or_cdn())
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"   TPM 趋势 HTML: {html_path}")


_HTML_TPM_TEMPLATE = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><title>TPM & QPM 趋势图</title>
__CHART_JS_TAG__
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#f5f7fa;color:#333;padding:20px}}
.container{{max-width:1280px;margin:0 auto}}
h1{{font-size:22px;margin-bottom:12px;color:#1a1a2e;border-left:4px solid #4361ee;padding-left:12px}}
.desc{{font-size:13px;line-height:1.7;color:#5b6472;margin-bottom:16px;background:#fff;border-radius:10px;padding:14px 18px;box-shadow:0 2px 12px rgba(0,0,0,.06)}}
.card{{background:#fff;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,.08);padding:22px;margin-bottom:20px}}
.card h2{{font-size:16px;margin-bottom:12px;color:#1f2937}}
.chart-wrap{{position:relative;height:420px;width:100%}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin-top:8px}}
th{{background:#f0f3ff;color:#4361ee;font-weight:600;padding:8px 12px;text-align:center;border:1px solid #e0e4ed;white-space:nowrap}}
td{{padding:6px 12px;text-align:right;border:1px solid #e8ecf2}}
td:first-child{{text-align:left;font-weight:500;background:#fafbff}}
tr:hover td{{background:#f8faff}}
.legend{{display:flex;gap:18px;justify-content:center;margin-bottom:12px;flex-wrap:wrap}}
.legend-item{{display:flex;align-items:center;gap:6px;font-size:13px}}
.dot{{width:12px;height:12px;border-radius:50%}}
.selector-bar{{display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}}
.selector-bar label{{font-size:15px;font-weight:600;color:#1a1a2e}}
.selector-bar select{{font-size:15px;padding:6px 16px;border-radius:8px;border:1px solid #d0d5e0;background:#fff;color:#333;cursor:pointer;outline:none}}
.selector-bar select:focus{{border-color:#4361ee;box-shadow:0 0 0 2px rgba(67,97,238,.18)}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
@media(max-width:980px){{.grid{{grid-template-columns:1fr}}}}
</style></head><body><div class="container">
<h1>性能趋势分析</h1>
<div class="desc">
  <b>口径：</b>Input TPM 按 <code>start_time</code> 归桶，近似请求注入 / prefill 压力；
  Output TPM 将 <code>completion_tokens</code> 均摊到 <code>start_time + TTFT</code> 到 <code>completed_time</code> 区间，近似流式生成吞吐；
  Total TPM = Input TPM + Output TPM；QPM 按成功请求 <code>completed_time</code> 归桶。默认 60s 滑动窗口，30s 步长。
</div>
<div class="selector-bar"><label for="labelSelect">数据档位：</label><select id="labelSelect">{options_html}</select></div>
<div class="card"><h2>Total TPM</h2><div class="legend" id="totalLegend"></div><div class="chart-wrap"><canvas id="totalChart"></canvas></div></div>
<div class="grid">
  <div class="card"><h2>Input TPM</h2><div class="legend" id="inputLegend"></div><div class="chart-wrap"><canvas id="inputChart"></canvas></div></div>
  <div class="card"><h2>Output TPM</h2><div class="legend" id="outputLegend"></div><div class="chart-wrap"><canvas id="outputChart"></canvas></div></div>
</div>
<div class="card"><h2>QPM</h2><div class="legend" id="qpmLegend"></div><div class="chart-wrap"><canvas id="qpmChart"></canvas></div></div>
<div class="card"><h2>趋势数据明细表</h2><table id="trendTable"><thead><tr></tr></thead><tbody></tbody></table></div>
</div>
<script>
window.addEventListener("load",function(){{
var ALL={labels_json};
var PALETTE={palette};
var charts={{}};
var METRICS={{
  total:{{chartId:"totalChart",legendId:"totalLegend",title:"Total TPM (tokens/min)"}},
  input:{{chartId:"inputChart",legendId:"inputLegend",title:"Input TPM (tokens/min)"}},
  output:{{chartId:"outputChart",legendId:"outputLegend",title:"Output TPM (tokens/min)"}},
  qpm:{{chartId:"qpmChart",legendId:"qpmLegend",title:"QPM (queries/min)"}}
}};
function buildSeries(RAW,cols){{
  var s={{}};
  RAW.forEach(function(r){{
    var t=r["时间(s)"];
    cols.forEach(function(k){{
      if(!s[k])s[k]=[];
      var v=r[k];
      if(v!==null&&v!==""&&v!==undefined)s[k].push({{x:t,y:Number(v)}});
    }});
  }});
  return s;
}}
function fmt(v){{return Math.abs(v)>=1000000?(v/1000000).toFixed(2)+"M":Math.abs(v)>=1000?(v/1000).toFixed(1)+"k":v;}}
function makeChart(metricKey,series){{
  var cfg=METRICS[metricKey];
  var keys=Object.keys(series).sort(function(a,b){{return parseInt((a.match(/P(\d+)/)||[])[1]||0)-parseInt((b.match(/P(\d+)/)||[])[1]||0);}});
  var leg=document.getElementById(cfg.legendId);leg.innerHTML="";
  keys.forEach(function(k,i){{
    var s=PALETTE[i%PALETTE.length];
    var d=document.createElement("div");d.className="legend-item";
    d.innerHTML='<span class="dot" style="background:'+s.border+'"></span>'+k;
    leg.appendChild(d);
  }});
  return new Chart(document.getElementById(cfg.chartId),{{
    type:"line",
    data:{{datasets:keys.map(function(k,i){{
      var s=PALETTE[i%PALETTE.length];
      return{{label:k,data:series[k],borderColor:s.border,backgroundColor:s.bg,fill:false,tension:.18,pointRadius:2.5,pointHoverRadius:7,pointBackgroundColor:s.border,pointBorderColor:"#fff",borderWidth:2.2}};
    }})}},
    options:{{
      responsive:true,maintainAspectRatio:false,interaction:{{mode:"index",intersect:false}},
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{title:function(i){{return"时间: "+i[0].parsed.x+"s"}},label:function(i){{return i.dataset.label+": "+Number(i.parsed.y).toLocaleString()}}}}}}}},
      scales:{{x:{{type:"linear",title:{{display:true,text:"时间 (s)"}},grid:{{color:"#eee"}},ticks:{{stepSize:30}}}},y:{{title:{{display:true,text:cfg.title}},grid:{{color:"#eee"}},ticks:{{callback:function(v){{return fmt(v)}}}}}}}}
    }}
  }});
}}
function makeTable(RAW,colsByMetric){{
  var thead=document.querySelector("#trendTable thead tr");thead.innerHTML="";
  var allCols=["时间(s)"];
  ["total","input","output","qpm"].forEach(function(k){{allCols=allCols.concat(colsByMetric[k]||[]);}});
  allCols.forEach(function(c){{var th=document.createElement("th");th.textContent=c;thead.appendChild(th);}});
  var tbody=document.querySelector("#trendTable tbody");tbody.innerHTML="";
  RAW.forEach(function(r){{
    var tr=document.createElement("tr");
    allCols.forEach(function(c){{
      var td=document.createElement("td");
      if(c==="时间(s)")td.textContent=r[c];
      else{{var v=r[c];td.textContent=(v!==null&&v!==undefined&&v!=="")?Number(v).toLocaleString():"-";}}
      tr.appendChild(td);
    }});
    tbody.appendChild(tr);
  }});
}}
function render(label){{
  var ds=ALL[label];if(!ds)return;
  Object.keys(charts).forEach(function(k){{if(charts[k]){{charts[k].destroy();charts[k]=null;}}}});
  ["total","input","output","qpm"].forEach(function(k){{
    var series=buildSeries(ds.rows,ds.cols[k]||[]);
    charts[k]=makeChart(k,series);
  }});
  makeTable(ds.rows,ds.cols);
}}
render("{first_label}");
document.getElementById("labelSelect").addEventListener("change",function(){{render(this.value);}});
}});
</script></body></html>'''


if __name__ == "__main__":
    main()
